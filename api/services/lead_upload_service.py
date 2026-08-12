import csv
import io
import logging
import re

from django.core.validators import validate_email
from django.core.exceptions import ValidationError

logger = logging.getLogger(__name__)

# Canonical Lead field -> accepted header aliases (case/space-insensitive).
_HEADER_ALIASES = {
    "email": ["email", "e-mail", "email address", "work email", "mail"],
    "first_name": ["first name", "firstname", "fname", "given name"],
    "last_name": ["last name", "lastname", "lname", "surname", "family name"],
    "person_full_name": ["full name", "name", "contact name", "person name"],
    "phone": ["phone", "phone number", "mobile", "contact number", "telephone"],
    "linkedin_url": ["linkedin", "linkedin url", "linkedin profile", "linkedin link"],
    "title": ["title", "job title", "position", "designation", "role"],
    "person_seniority": ["seniority"],
    "person_department": ["department"],
    "company_name": ["company", "company name", "organization", "organisation", "employer"],
    "company_domain": ["domain", "company domain", "website"],
    "industry": ["industry"],
    "location": ["location", "city", "address"],
}

_MAX_ROWS = 5000


def _normalize_header(raw: str) -> str:
    return re.sub(r"\s+", " ", (raw or "").strip().lower())


def _build_column_map(headers: list) -> dict:
    """Maps column index -> canonical Lead field name, based on header text."""
    alias_lookup = {}
    for canonical, aliases in _HEADER_ALIASES.items():
        for alias in aliases:
            alias_lookup[alias] = canonical

    column_map = {}
    for idx, raw_header in enumerate(headers):
        normalized = _normalize_header(str(raw_header) if raw_header is not None else "")
        canonical = alias_lookup.get(normalized)
        if canonical and canonical not in column_map.values():
            column_map[idx] = canonical
    return column_map


def _rows_from_csv(file_obj):
    text = file_obj.read().decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return [], []
    return rows[0], rows[1:]


def _rows_from_xlsx(file_obj):
    from openpyxl import load_workbook

    wb = load_workbook(file_obj, read_only=True, data_only=True)
    sheet = wb.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header = list(next(rows_iter))
    except StopIteration:
        return [], []
    data_rows = [list(row) for row in rows_iter]
    return header, data_rows


def parse_lead_file(file_obj, filename: str):
    """
    Parses an uploaded .csv or .xlsx file into a list of raw lead dicts,
    using flexible header matching. Returns (records, parse_errors).
    """
    lower_name = (filename or "").lower()
    if lower_name.endswith(".csv"):
        header, data_rows = _rows_from_csv(file_obj)
    elif lower_name.endswith(".xlsx"):
        header, data_rows = _rows_from_xlsx(file_obj)
    else:
        raise ValueError("Unsupported file type. Please upload a .csv or .xlsx file.")

    if not header:
        return [], ["File is empty or has no header row."]

    column_map = _build_column_map(header)
    if "email" not in column_map.values():
        raise ValueError(
            "No 'email' column found. The file must include an Email column "
            "(accepted headers: " + ", ".join(_HEADER_ALIASES["email"]) + ")."
        )

    records = []
    for row in data_rows:
        if row is None or all(cell in (None, "") for cell in row):
            continue
        record = {}
        for idx, canonical in column_map.items():
            if idx < len(row) and row[idx] not in (None, ""):
                record[canonical] = str(row[idx]).strip()
        records.append(record)

    return records, []


def import_leads(file_obj, filename: str, lead_list_name: str | None = None, owner_user_id: str | None = None) -> dict:
    """
    Parses an uploaded Excel/CSV file and upserts rows into the Lead table
    (channel="upload"), deduplicated on email. Optionally groups the
    imported leads into a new LeadList.
    """
    from ..models import Lead, LeadList

    records, parse_errors = parse_lead_file(file_obj, filename)
    if parse_errors:
        return {"imported": 0, "updated": 0, "skipped": 0, "errors": parse_errors, "total_rows": 0}

    if len(records) > _MAX_ROWS:
        raise ValueError(f"File has {len(records)} rows; max {_MAX_ROWS} rows per upload.")

    lead_list = None
    if lead_list_name:
        lead_list = LeadList.objects.create(name=lead_list_name)

    created_count = 0
    updated_count = 0
    skipped_count = 0
    row_errors = []
    seen_emails_in_file = set()

    for row_num, record in enumerate(records, start=2):  # +1 header, +1 to be 1-indexed
        email = (record.get("email") or "").strip().lower()

        if not email:
            skipped_count += 1
            row_errors.append({"row": row_num, "reason": "Missing email"})
            continue

        try:
            validate_email(email)
        except ValidationError:
            skipped_count += 1
            row_errors.append({"row": row_num, "reason": f"Invalid email: {email}"})
            continue

        if email in seen_emails_in_file:
            skipped_count += 1
            row_errors.append({"row": row_num, "reason": f"Duplicate email in file: {email}"})
            continue
        seen_emails_in_file.add(email)

        defaults = {
            "channel": "upload",
            "first_name": record.get("first_name"),
            "last_name": record.get("last_name"),
            "person_full_name": record.get("person_full_name"),
            "phone": record.get("phone"),
            "linkedin_url": record.get("linkedin_url"),
            "title": record.get("title"),
            "person_seniority": record.get("person_seniority"),
            "person_department": record.get("person_department"),
            "company_name": record.get("company_name"),
            "company_domain": record.get("company_domain"),
            "industry": record.get("industry"),
            "location": record.get("location"),
        }
        # Drop empty values so we don't overwrite existing enriched data with blanks.
        defaults = {k: v for k, v in defaults.items() if v not in (None, "")}
        defaults["channel"] = "upload"
        if lead_list is not None:
            defaults["lead_list"] = lead_list
        if owner_user_id:
            defaults["owner_user_id"] = owner_user_id

        try:
            lead, was_created = Lead.objects.update_or_create(email=email, defaults=defaults)
        except Exception as e:
            logger.error(f"Failed to upsert lead row {row_num} ({email}): {e}", exc_info=True)
            skipped_count += 1
            row_errors.append({"row": row_num, "reason": "Database error"})
            continue

        if was_created:
            created_count += 1
        else:
            updated_count += 1

    result = {
        "imported": created_count,
        "updated": updated_count,
        "skipped": skipped_count,
        "errors": row_errors,
        "total_rows": len(records),
    }
    if lead_list is not None:
        result["lead_list_id"] = lead_list.id
        result["lead_list_name"] = lead_list.name

    logger.info(
        f"Lead upload complete: created={created_count}, updated={updated_count}, "
        f"skipped={skipped_count}, total_rows={len(records)}"
    )
    return result
